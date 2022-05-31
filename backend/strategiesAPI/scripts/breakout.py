from alice_blue import *
from time import sleep
import dateutil.parser
import datetime
import pandas as pd
import pandasql as ps
from pyparsing import col
import requests
import openpyxl

instrument_list = [
    'ACC', 'AUBANK', 'ADANIENT', 'ADANIPORTS', 'AMBUJACEM', 'APOLLOHOSP',
    'ASIANPAINT', 'AUROPHARMA', 'AXISBANK', 'BAJAJ-AUTO', 'BAJFINANCE',
    'BATAINDIA', 'BHARATFORG', 'BPCL', 'BHARTIARTL', 'BIOCON', 'CHOLAFIN',
    'CIPLA', 'COALINDIA', 'COFORGE', 'DLF', 'DABUR', 'DIVISLAB', 'DRREDDY',
    'EICHERMOT', 'GODREJCP', 'GODREJPROP', 'GRASIM', 'HCLTECH', 'HDFCBANK',
    'HDFCLIFE', 'HAVELLS', 'HEROMOTOCO', 'HINDALCO', 'HINDPETRO', 'HINDUNILVR',
    'HDFC', 'ICICIBANK', 'ICICIPRULI', 'ITC', 'IRCTC', 'IGL', 'INDUSINDBK',
    'INFY', 'INDIGO', 'JSWSTEEL', 'JINDALSTEL', 'JUBLFOOD', 'KOTAKBANK',
    'LICHSGFIN', 'LTI', 'LT', 'LUPIN', 'M&M', 'MANAPPURAM', 'MARUTI',
    'MINDTREE', 'MUTHOOTFIN', 'PVR', 'PIDILITIND', 'PEL', 'RELIANCE',
    'SBICARD', 'SBILIFE', 'SRF', 'SRTRANSFIN', 'SBIN', 'SUNPHARMA', 'TVSMOTOR',
    'TATACHEM', 'TCS', 'TATACONSUM', 'TATAMOTORS', 'TATAPOWER', 'TATASTEEL',
    'TECHM', 'TITAN', 'UPL', 'VEDL', 'VOLTAS', 'WIPRO', 'ZEEL'
]

access_token = AliceBlue.login_and_get_access_token(username=username,
                                                    password=password,
                                                    twoFA=twoFA,
                                                    api_secret=api_secret,
                                                    app_id=app_id)
alice = AliceBlue(username=username,
                  password=password,
                  access_token=access_token)
traded_stocks = []


def main():
    # while ((datetime.datetime.now().time() < datetime.time(9, 20, 00))):
    #     pass
    wrkbk = openpyxl.load_workbook(
        f'{datetime.datetime.now().strftime("%Y-%m-%d")}.xlsx')
    sh = wrkbk.active

    x = 2 + len(instrument_list)
    for i in instrument_list:
        name = sh.cell(row=x, column=3).value
        vo = test(name)
        vol = sh.cell(row=x, column=4).value
        open = sh.cell(row=x, column=5).value
        close = sh.cell(row=x, column=6).value
        high = sh.cell(row=x, column=7).value
        low = sh.cell(row=x, column=8).value
        print(f'{name} {vol} {vo}')
        p_open = vo[1] * 0.06
        p_diff_b = p_open + vo[1]
        p_diff_s = vo[1] - p_open
        c_high = close - close * 0.03
        c_low = close + close * 0.03
        range_OC2 = close - open
        range_HL2 = high - low
        range_HL1 = high - low
        range_OC1 = open - close
        money = 100000
        L1 = low - low * 0.1 / 100
        H1 = high + high * 0.1 / 100
        quantity_b = int(money / H1)
        quantity_s = int(money / L1)

        if ((vol > vo[0]) and (open < p_diff_b) and (close < c_low)
                and (range_OC2 > range_HL2 * 0.80)):
            print(f"Entry {name}")
            alice.place_order(transaction_type=TransactionType.Buy,
                              instrument=alice.get_instrument_by_symbol(
                                  'NSE', name),
                              quantity=quantity_b,
                              order_type=OrderType.Market,
                              product_type=ProductType.Delivery,
                              price=high,
                              trigger_price=None,
                              stop_loss=low,
                              square_off=2 * (high - low),
                              trailing_sl=None,
                              is_amo=False)

        if ((vol > vo[0]) and (open > p_diff_s) and (close > c_high)
                and (range_OC1 > range_HL1 * 0.80)):
            print(f"Exit {name}")
            alice.place_order(transaction_type=TransactionType.Sell,
                              instrument=alice.get_instrument_by_symbol(
                                  'NSE', name),
                              quantity=quantity_s,
                              order_type=OrderType.Market,
                              product_type=ProductType.Intraday,
                              price=low,
                              trigger_price=None,
                              stop_loss=low,
                              square_off=2 * (high - low),
                              trailing_sl=None,
                              is_amo=False)
        x += 1


def get_historical(instrument,
                   from_datetime,
                   to_datetime,
                   interval,
                   indices=False):
    params = {
        "token":
        instrument.token,
        "exchange":
        instrument.exchange if not indices else "NSE_INDICES",
        "starttime":
        str(int(from_datetime.timestamp())),
        "endtime":
        str(int(to_datetime.timestamp())),
        "candletype":
        3 if interval.upper() == "DAY" else
        (2 if interval.upper().split("_")[1] == "HR" else 1),
        "data_duration":
        None if interval.upper() == "DAY" else interval.split("_")[0]
    }
    lst = requests.get(f" https://ant.aliceblueonline.com/api/v1/charts/tdv?",
                       params=params).json()["data"]["candles"]
    records = []
    for i in lst:
        record = {
            "date": dateutil.parser.parse(i[0]),
            "open": i[1],
            "high": i[2],
            "low": i[3],
            "close": i[4],
            "volume": i[5]
        }
        records.append(record)
    return records


def test(i):
    instrument = alice.get_instrument_by_symbol("NSE", i)
    date = datetime.datetime.strptime('2017-05-04', "%Y-%m-%d")
    date_start = date.replace(minute=15,
                              hour=9,
                              second=00,
                              year=datetime.datetime.now().year,
                              month=datetime.datetime.now().month,
                              day=datetime.datetime.now().day)
    date_end = date.replace(minute=30,
                            hour=15,
                            second=00,
                            year=datetime.datetime.now().year,
                            month=datetime.datetime.now().month,
                            day=datetime.datetime.now().day)
    from_datetime = date_start - datetime.timedelta(
        days=1)  ##[ON Monday Value Need to change to 4]
    to_datetime = date_end - datetime.timedelta(
        days=1)  ##[ON Monday Value Need to change to 3]
    interval1 = "5_MIN"  # ["DAY", "1_HR", "3_HR", "1_MIN", "5_MIN", "15_MIN", "60_MIN"]
    interval2 = "DAY"  # ["DAY", "1_HR", "3_HR", "1_MIN", "5_MIN", "15_MIN", "60_MIN"]
    indices = False
    df1 = pd.DataFrame(
        get_historical(instrument, from_datetime, to_datetime, interval1,
                       indices))

    df2 = pd.DataFrame(
        get_historical(instrument, from_datetime, to_datetime, interval2,
                       indices))

    q1 = """SELECT  volume FROM df1 order by volume DESC LIMIT 1 """
    q2 = """SELECT  open FROM df2 """

    s1 = ps.sqldf(q1, locals())
    s2 = ps.sqldf(q2, locals())

    vol = s1["volume"][0]
    op = s2["open"][0]
    return (vol, op)
