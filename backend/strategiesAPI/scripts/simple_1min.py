import time
from alice_blue import *
import dateutil.parser
import datetime
import pandas as pd
import pandasql as ps
import requests
import openpyxl
import re
import django

django.setup()
from ..models import Papertrade, TradedStocks

traded_stocks = []
instrument_list = [
    'ACC', 'ADANIENT', 'ADANIPORTS', 'AMBUJACEM', 'APOLLOHOSP', 'ASIANPAINT', 'AUBANK',
    'AUROPHARMA', 'AXISBANK', 'BAJAJ-AUTO', 'BAJFINANCE', 'BATAINDIA', 'BHARATFORG',
    'BHARTIARTL', 'BIOCON', 'BPCL', 'CHOLAFIN', 'CIPLA', 'COALINDIA', 'COFORGE', 'DABUR',
    'DIVISLAB', 'DLF', 'DRREDDY', 'EICHERMOT', 'GODREJCP', 'GODREJPROP', 'GRASIM', 'HAVELLS',
    'HCLTECH', 'HDFC', 'HDFCBANK', 'HDFCLIFE', 'HEROMOTOCO', 'HINDALCO', 'HINDPETRO',
    'HINDUNILVR', 'ICICIBANK', 'ICICIPRULI', 'IGL', 'INDIGO', 'INDUSINDBK', 'INFY',
    'IRCTC', 'ITC', 'JINDALSTEL', 'JSWSTEEL', 'JUBLFOOD', 'KOTAKBANK', 'LICHSGFIN',
    'LT', 'LTI', 'LUPIN', 'M&M', 'MANAPPURAM', 'MARUTI', 'MINDTREE', 'MUTHOOTFIN',
    'PEL', 'PIDILITIND', 'PVR', 'RELIANCE', 'SBICARD', 'SBILIFE', 'SBIN', 'SRF',
    'SRTRANSFIN', 'SUNPHARMA', 'TATACHEM', 'TATACONSUM', 'TATAMOTORS', 'TATAPOWER',
    'TATASTEEL', 'TCS', 'TECHM', 'TITAN', 'TVSMOTOR', 'UPL', 'VEDL', 'VOLTAS', 'WIPRO',
    'ZEEL'
]

access_token = AliceBlue.login_and_get_access_token(username=username,
                                                    password=password,
                                                    twoFA=twoFA,
                                                    api_secret=api_secret,
                                                    app_id=app_id)
alice = AliceBlue(username=username,
                  password=password,
                  access_token=access_token)

df_historical = pd.DataFrame()


def get_historical(instrument,
                   from_datetime,
                   to_datetime,
                   interval,
                   indices=False):
    params = {
        "token":
            instrument.token,
        "exchange":
            instrument.exchange if not indices else "NSE_INDICES",
        "starttime":
            str(int(from_datetime.timestamp())),
        "endtime":
            str(int(to_datetime.timestamp())),
        "candletype":
            3 if interval.upper() == "DAY" else
            (2 if interval.upper().split("_")[1] == "HR" else 1),
        "data_duration":
            None if interval.upper() == "DAY" else interval.split("_")[0]
    }
    lst = requests.get(f" https://ant.aliceblueonline.com/api/v1/charts/tdv?",
                       params=params).json()["data"]["candles"]
    records = []
    for i in lst:
        record = {
            "date": dateutil.parser.parse(i[0]),
            "open": i[1],
            "high": i[2],
            "low": i[3],
            "close": i[4],
            "volume": i[5]
        }
        records.append(record)
    return records


def test(i):
    instrument = alice.get_instrument_by_symbol("NSE", i)
    date = datetime.datetime.strptime('2017-05-04', "%Y-%m-%d")
    date_start = date.replace(minute=15,
                              hour=9,
                              second=00,
                              year=datetime.datetime.now().year,
                              month=datetime.datetime.now().month,
                              day=datetime.datetime.now().day)
    date_end = date.replace(minute=30,
                            hour=15,
                            second=00,
                            year=datetime.datetime.now().year,
                            month=datetime.datetime.now().month,
                            day=datetime.datetime.now().day)

    # If weekday is Monday, historical will be of friday
    if datetime.datetime.now().weekday() == 0:
        from_datetime = date_start - datetime.timedelta(
            days=3)
        to_datetime = date_end - datetime.timedelta(
            days=3)
    else:
        from_datetime = date_start - datetime.timedelta(
            days=1)
        to_datetime = date_end - datetime.timedelta(
            days=1)

    interval1 = "1_MIN"  # ["DAY", "1_HR", "3_HR", "1_MIN", "5_MIN", "15_MIN", "60_MIN"]
    indices = False
    df1 = pd.DataFrame(
        get_historical(instrument, from_datetime, to_datetime, interval1,
                       indices))

    q1 = """SELECT  volume FROM df1 order by volume DESC LIMIT 1 """
    s1 = ps.sqldf(q1, locals())
    vol = s1["volume"][0]

    return vol


def row_count(ws):
    return int(re.search('(\d+)$', ws.dimensions).group(1))


def main():
    global df_historical
    for x in range(len(instrument_list)):
        name = instrument_list[x]
        vo = test(name)
        data = {'symbol': name, 'volume': vo}
        df_histo = pd.DataFrame(data, index=[x])
        df_historical = pd.concat([df_historical, df_histo])
        print(f"{df_historical['symbol'][x]}  {df_historical['volume'][x]} {x}")

    while datetime.datetime.now().time() < datetime.time(9, 15, 00):
        pass
    interval = 60 - datetime.datetime.now().second
    time.sleep(interval + 2)

    while datetime.time(9, 16, 2) <= datetime.datetime.now().time() <= datetime.time(15, 25, 5):
        start = time.time()
        wrkbk = openpyxl.load_workbook(
            f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}_1MIN.xlsx')
        wrkbk.active = wrkbk['Sheet1']
        sh = wrkbk.active
        rows_num = row_count(sh)
        print(f'1MIN candle at {datetime.datetime.now().time().strftime("%H:%M")}')
        x = rows_num - 82
        for y in range(len(instrument_list)):
            name = sh.cell(row=x, column=2).value
            if name == 'BANKNIFTY JUL FUT':
                y -= 1
                x += 1
                continue

            vo = df_historical['volume'][y]
            vol = sh.cell(row=x, column=3).value
            open = sh.cell(row=x, column=4).value
            close = sh.cell(row=x, column=5).value
            high = sh.cell(row=x, column=6).value
            low = sh.cell(row=x, column=7).value
            atp = sh.cell(row=x, column=9).value
            range_oc2 = close - open
            range_hl2 = high - low
            range_hl1 = high - low
            range_oc1 = open - close
            money = 100000
            l1 = low - low * 0.1 / 100
            h1 = high + high * 0.1 / 100
            quantity_b = int(money / h1)
            quantity_s = int(money / l1)

            if ((name not in traded_stocks) and (vol > vo) and (open < close)
                    and (range_oc2 > range_hl2 * 0.80) and (close > atp)):
                traded_stocks.append(name)
                print(f"Entry {name} {vol} {vo}")
                alice.place_order(transaction_type=TransactionType.Buy,
                                  instrument=alice.get_instrument_by_symbol(
                                      'NSE', name),
                                  quantity=quantity_b,
                                  order_type=OrderType.Market,
                                  product_type=ProductType.Delivery,
                                  price=0.0,
                                  trigger_price=None,
                                  stop_loss=float(low),
                                  square_off=float(2 * (high - low)),
                                  trailing_sl=None,
                                  is_amo=False)

            if (name not in traded_stocks) and (vol > vo) and (open > close) and (range_oc1 > range_hl1 * 0.80) and (
                    close < atp):
                print(f"Exit {name} {vol} {vo}")
                traded_stocks.append(name)
                alice.place_order(transaction_type=TransactionType.Sell,
                                  instrument=alice.get_instrument_by_symbol(
                                      'NSE', name),
                                  quantity=quantity_s,
                                  order_type=OrderType.Market,
                                  product_type=ProductType.Intraday,
                                  price=0.0,
                                  trigger_price=None,
                                  stop_loss=float(low),
                                  square_off=float(2 * (high - low)),
                                  trailing_sl=None,
                                  is_amo=False)
            x += 1

            wrkbk.close()
        interval = 60 - (time.time() - start)
        time.sleep(interval)


def start_paper_trade():
    global df_historical
    # noinspection PyGlobalUndefined
    global algotrade_username
    for x in range(len(instrument_list)):
        name = instrument_list[x]
        vo = test(name)
        data = {'symbol': name, 'volume': vo}
        df_histo = pd.DataFrame(data, index=[x])
        df_historical = pd.concat([df_historical, df_histo])
        print(f"{df_historical['symbol'][x]}  {df_historical['volume'][x]} {x}")

    while datetime.datetime.now().time() < datetime.time(9, 15, 00):
        pass

    interval = 60 - datetime.datetime.now().second
    time.sleep(interval + 2)

    while datetime.time(9, 16, 0) <= datetime.datetime.now().time() <= datetime.time(15, 0, 0):
        start = time.time()
        wrkbk = openpyxl.load_workbook(
            f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}_1MIN.xlsx')
        wrkbk.active = wrkbk['Sheet1']
        sh = wrkbk.active
        rows_num = row_count(sh)
        x = rows_num - 82
        print(f'1MIN candle at {datetime.datetime.now().time().strftime("%H:%M")}')
        for y in range(len(instrument_list)):
            name = sh.cell(row=x, column=2).value
            if name == 'BANKNIFTY JUL FUT':
                y -= 1
                x += 1
                continue

            vo = df_historical['volume'][y]
            vol = sh.cell(row=x, column=3).value
            open = sh.cell(row=x, column=4).value
            close = sh.cell(row=x, column=5).value
            high = sh.cell(row=x, column=6).value
            low = sh.cell(row=x, column=7).value
            atp = sh.cell(row=x, column=9).value
            range_oc2 = close - open
            range_hl2 = high - low
            range_hl1 = high - low
            range_oc1 = open - close
            money = 100000
            l1 = low - low * 0.1 / 100
            h1 = high + high * 0.1 / 100
            quantity_b = int(money / h1)
            quantity_s = int(money / l1)
            if ((not TradedStocks.objects.filter(username=algotrade_username).filter(stock_name=name).exists()) and (
                    vol > vo) and (open < close)
                    and (range_oc2 > range_hl2 * 0.90) and (close > atp)):
                print(f"Entry {name} {vol} {vo}")
                stop_loss = high + 0.5 - (high - low) / 2
                square_off = (high - low) / 2
                TradedStocks.objects.create(username=algotrade_username, stock_name=name)
                Papertrade.objects.create(start_time=datetime.datetime.now().time().strftime("%H:%M"),
                                          username=algotrade_username, signal='BUY', name=name, quantity=quantity_b,
                                          buy_price=high + 0.5, sell_price=0, stop_loss=stop_loss, target=square_off)

            if ((not TradedStocks.objects.filter(username=algotrade_username).filter(stock_name=name).exists()) and (
                    vol > vo) and (open > close) and (range_oc1 > range_hl1 * 0.90) and (close < atp)):
                print(f"Exit {name} {vol} {vo}")
                stop_loss = low - 0.5 + (high - low) / 2
                square_off = (high - low) / 2
                TradedStocks.objects.create(username=algotrade_username, stock_name=name)
                Papertrade.objects.create(start_time=datetime.datetime.now().time().strftime("%H:%M"),
                                          username=algotrade_username, signal='SELL', name=name, quantity=quantity_s,
                                          buy_price=0, sell_price=low - 0.5, stop_loss=stop_loss, target=square_off)

            x += 1
            wrkbk.close()
        interval = 60 - time.time() + start
        time.sleep(interval)
