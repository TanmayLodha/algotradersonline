import time
from alice_blue import *
import dateutil.parser
import datetime
import pandas as pd
import pandasql as ps
import requests
import openpyxl
import re
import django

django.setup()
from ..models import Papertrade

traded_stocks = []
instrument_list = ['BANKNIFTY JUL FUT', 'NIFTY JUL FUT']

access_token = AliceBlue.login_and_get_access_token(username=username,
                                                    password=password,
                                                    twoFA=twoFA,
                                                    api_secret=api_secret,
                                                    app_id=app_id)
alice = AliceBlue(username=username,
                  password=password,
                  access_token=access_token)

df_historical = pd.DataFrame()


def get_historical(instrument,
                   from_datetime,
                   to_datetime,
                   interval,
                   indices=False):
    params = {
        "token":
            instrument.token,
        "exchange":
            instrument.exchange if not indices else "NSE_INDICES",
        "starttime":
            str(int(from_datetime.timestamp())),
        "endtime":
            str(int(to_datetime.timestamp())),
        "candletype":
            3 if interval.upper() == "DAY" else
            (2 if interval.upper().split("_")[1] == "HR" else 1),
        "data_duration":
            None if interval.upper() == "DAY" else interval.split("_")[0]
    }
    lst = requests.get(f" https://ant.aliceblueonline.com/api/v1/charts/tdv?",
                       params=params).json()["data"]["candles"]
    records = []
    for i in lst:
        record = {
            "date": dateutil.parser.parse(i[0]),
            "open": i[1],
            "high": i[2],
            "low": i[3],
            "close": i[4],
            "volume": i[5]
        }
        records.append(record)
    return records


def test(i):
    instrument = alice.get_instrument_by_symbol("NFO", i)
    date = datetime.datetime.strptime('2017-05-04', "%Y-%m-%d")
    date_start = date.replace(minute=15,
                              hour=9,
                              second=00,
                              year=datetime.datetime.now().year,
                              month=datetime.datetime.now().month,
                              day=datetime.datetime.now().day)
    date_end = date.replace(minute=30,
                            hour=9,
                            second=00,
                            year=datetime.datetime.now().year,
                            month=datetime.datetime.now().month,
                            day=datetime.datetime.now().day)
    from_datetime = date_start - datetime.timedelta(
        days=0)
    to_datetime = date_end - datetime.timedelta(
        days=0)
    interval1 = "15_MIN"  # ["DAY", "1_HR", "3_HR", "1_MIN", "5_MIN", "15_MIN", "60_MIN"]
    indices = False
    df1 = pd.DataFrame(
        get_historical(instrument, from_datetime, to_datetime, interval1,
                       indices))

    q1 = """SELECT high FROM df1 order by high DESC LIMIT 1 """
    q2 = """SELECT low FROM df1 order by low ASC LIMIT 1 """
    s1 = ps.sqldf(q1, locals())
    s2 = ps.sqldf(q2, locals())

    high = s1["high"][0]
    low = s2["low"][0]

    return high, low


def row_count(ws):
    return int(re.search('(\d+)$', ws.dimensions).group(1))


def start_paper_trade():
    global df_historical
    # noinspection PyGlobalUndefined
    global algotrade_username
    for x in range(len(instrument_list)):
        name = instrument_list[x]
        high, low = test(name)
        data = {'symbol': name, 'high': high, 'low': low}
        df_histo = pd.DataFrame(data, index=[x])
        df_historical = pd.concat([df_historical, df_histo])
        print(f"{df_historical['symbol'][x]}  {df_historical['high'][x]} {df_historical['low'][x]}")

    while datetime.datetime.now().time() < datetime.time(9, 30, 00):
        pass

    interval = 60 - datetime.datetime.now().second
    time.sleep(interval + 2)

    while datetime.time(9, 31, 0) <= datetime.datetime.now().time() <= datetime.time(15, 0, 0):
        start = time.time()
        wrkbk = openpyxl.load_workbook(
            f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}_1MIN.xlsx')
        wrkbk.active = wrkbk['Sheet1']
        sh = wrkbk.active
        rows_num = row_count(sh)

        print(f'FUTURES at {datetime.datetime.now().time().strftime("%H:%M")}')
        for x in range(rows_num - 84, rows_num):
            name = sh.cell(row=x, column=2).value
            y = 0 if name == "BANKNIFTY JUL FUT" else 1
            if name == "BANKNIFTY JUL FUT" or name == 'NIFTY JUL FUT':
                hi = df_historical['high'][y]
                lo = df_historical['low'][y]
                close = sh.cell(row=x, column=5).value
                high = sh.cell(row=x, column=6).value
                low = sh.cell(row=x, column=7).value
                if (name not in traded_stocks) and close > hi:
                    traded_stocks.append(name)
                    if name == "BANKNIFTY JUL FUT":
                        # square_off=100
                        stop_loss = hi - 50
                    else:
                        # square_off=50
                        stop_loss = hi - 30
                    print(f"Entry {name} {high} {hi}")
                    Papertrade.objects.create(start_time=datetime.datetime.now().time().strftime("%H:%M"),
                                              username=algotrade_username, signal='BUY', name=name, quantity=50,
                                              buy_price=hi + 20, sell_price=hi + 20, stop_loss=stop_loss, sl_trail=True,
                                              Invested=(hi + 20) * 50)

                if (name not in traded_stocks) and close < lo:
                    print(f"Exit {name} {low} {lo}")
                    traded_stocks.append(name)
                    if name == "BANKNIFTY JUL FUT":
                        # square_off = 100
                        stop_loss = lo + 50
                    else:
                        # square_off=50
                        stop_loss = lo + 30
                    Papertrade.objects.create(start_time=datetime.datetime.now().time().strftime("%H:%M"),
                                              username=algotrade_username, signal='SELL', name=name, quantity=50,
                                              buy_price=lo - 20, sell_price=lo - 20, stop_loss=stop_loss, sl_trail=True,
                                              Invested=(lo - 20) * 50)

        wrkbk.close()
        interval = 60 - time.time() + start
        time.sleep(interval)
