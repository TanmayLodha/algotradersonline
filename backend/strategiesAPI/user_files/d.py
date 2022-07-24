# config
username = "411686"
password = "ajeet302031"
app_id = "SX5Vs0UiCE"
api_secret = "NfL6k8K82kryxfAusAtUOsLUdJzSY7TjrHvMmrsO22WHiRTOuc4QCKGMw27C1uly"
twoFA = "1979"
algotrade_username = "d"

from alice_blue import *
import dateutil.parser
import datetime
import pandas as pd
import pandasql as ps
import requests
import openpyxl
import re

instrument_list = [
    'ACC', 'ADANIENT', 'ADANIPORTS', 'AMBUJACEM', 'APOLLOHOSP', 'ASIANPAINT', 'AUBANK',
    'AUROPHARMA', 'AXISBANK', 'BAJAJ-AUTO', 'BAJFINANCE', 'BATAINDIA', 'BHARATFORG',
    'BHARTIARTL', 'BIOCON', 'BPCL', 'CHOLAFIN', 'CIPLA', 'COALINDIA', 'COFORGE', 'DABUR',
    'DIVISLAB', 'DLF', 'DRREDDY', 'EICHERMOT', 'GODREJCP', 'GODREJPROP', 'GRASIM', 'HAVELLS',
    'HCLTECH', 'HDFC', 'HDFCBANK', 'HDFCLIFE', 'HEROMOTOCO', 'HINDALCO', 'HINDPETRO',
    'HINDUNILVR', 'ICICIBANK', 'ICICIPRULI', 'IGL', 'INDIGO', 'INDUSINDBK', 'INFY',
    'IRCTC', 'ITC', 'JINDALSTEL', 'JSWSTEEL', 'JUBLFOOD', 'KOTAKBANK', 'LICHSGFIN',
    'LT', 'LTI', 'LUPIN', 'M&M', 'MANAPPURAM', 'MARUTI', 'MINDTREE', 'MUTHOOTFIN',
    'PEL', 'PIDILITIND', 'PVR', 'RELIANCE', 'SBICARD', 'SBILIFE', 'SBIN', 'SRF',
    'SRTRANSFIN', 'SUNPHARMA', 'TATACHEM', 'TATACONSUM', 'TATAMOTORS', 'TATAPOWER',
    'TATASTEEL', 'TCS', 'TECHM', 'TITAN', 'TVSMOTOR', 'UPL', 'VEDL', 'VOLTAS', 'WIPRO',
    'ZEEL'
]

access_token = AliceBlue.login_and_get_access_token(username=username,
                                                    password=password,
                                                    twoFA=twoFA,
                                                    api_secret=api_secret,
                                                    app_id=app_id)
alice = AliceBlue(username=username,
                  password=password,
                  access_token=access_token)
traded_stocks = []
df_historical = pd.DataFrame()


def row_count(ws):
    return int(re.search('(\d+)$', ws.dimensions).group(1))


def main():
    global df_historical
    for x in range(len(instrument_list)):
        name = instrument_list[x]
        vo = test(name)
        data = {'symbol': name, 'volume': vo[0], 'open': vo[1], 'close': vo[2]}
        df_histo = pd.DataFrame(data, index=[x])
        df_historical = pd.concat([df_historical, df_histo])

    while datetime.datetime.now().time() < datetime.time(9, 20, 5):
        pass
    wrkbk = openpyxl.load_workbook(
        f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}.xlsx')
    wrkbk.active = wrkbk['Sheet1']
    sh = wrkbk.active
    rows_num = row_count(sh)
    print(rows_num)
    x = rows_num - len(instrument_list) + 1
    for y in range(len(instrument_list)):
        name = sh.cell(row=x, column=2).value
        vo = df_historical['volume'][y]
        op = df_historical['open'][y]
        clo = df_historical['close'][y]
        vol = sh.cell(row=x, column=3).value
        open = sh.cell(row=x, column=4).value
        close = sh.cell(row=x, column=5).value
        high = sh.cell(row=x, column=6).value
        low = sh.cell(row=x, column=7).value
        print(f'{name} {vol} {vo}')
        p_open = op * 0.06
        p_diff_b = p_open + op
        p_diff_s = op - p_open
        c_high = clo - clo * 0.03
        c_low = clo + clo * 0.03
        range_oc2 = close - open
        range_hl2 = high - low
        range_hl1 = high - low
        range_oc1 = open - close
        money = 100000
        l1 = low - low * 0.1 / 100
        h1 = high + high * 0.1 / 100
        quantity_b = int(money / h1)
        quantity_s = int(money / l1)

        if (vol > vo) and (open < p_diff_b) and (range_oc2 > range_hl2 * 0.80) and (open <= c_high):
            print(f"Entry {name}")
            alice.place_order(transaction_type=TransactionType.Buy,
                              instrument=alice.get_instrument_by_symbol(
                                  'NSE', name),
                              quantity=quantity_b,
                              order_type=OrderType.Market,
                              product_type=ProductType.Delivery,
                              price=0.0,
                              trigger_price=None,
                              stop_loss=None,
                              square_off=2 * (high - low),
                              trailing_sl=None,
                              is_amo=False)
        if (vol > vo) and (open > p_diff_s) and (range_oc1 > range_hl1 * 0.80) and (open <= c_low):
            print(f"Exit {name}")
            alice.place_order(transaction_type=TransactionType.Sell,
                              instrument=alice.get_instrument_by_symbol(
                                  'NSE', name),
                              quantity=quantity_s,
                              order_type=OrderType.Market,
                              product_type=ProductType.Intraday,
                              price=0.0,
                              trigger_price=None,
                              stop_loss=None,
                              square_off=2 * (high - low),
                              trailing_sl=None,
                              is_amo=False)

        x += 1


def get_historical(instrument,
                   from_datetime,
                   to_datetime,
                   interval,
                   indices=False):
    params = {
        "token":
            instrument.token,
        "exchange":
            instrument.exchange if not indices else "NSE_INDICES",
        "starttime":
            str(int(from_datetime.timestamp())),
        "endtime":
            str(int(to_datetime.timestamp())),
        "candletype":
            3 if interval.upper() == "DAY" else
            (2 if interval.upper().split("_")[1] == "HR" else 1),
        "data_duration":
            None if interval.upper() == "DAY" else interval.split("_")[0]
    }
    lst = requests.get(f" https://ant.aliceblueonline.com/api/v1/charts/tdv?",
                       params=params).json()["data"]["candles"]
    records = []
    for i in lst:
        record = {
            "date": dateutil.parser.parse(i[0]),
            "open": i[1],
            "high": i[2],
            "low": i[3],
            "close": i[4],
            "volume": i[5]
        }
        records.append(record)
    return records


def test(i):
    instrument = alice.get_instrument_by_symbol("NSE", i)
    date = datetime.datetime.strptime('2017-05-04', "%Y-%m-%d")
    date_start = date.replace(minute=15,
                              hour=9,
                              second=00,
                              year=datetime.datetime.now().year,
                              month=datetime.datetime.now().month,
                              day=datetime.datetime.now().day)
    date_end = date.replace(minute=30,
                            hour=15,
                            second=00,
                            year=datetime.datetime.now().year,
                            month=datetime.datetime.now().month,
                            day=datetime.datetime.now().day)
    from_datetime = date_start - datetime.timedelta(
        days=1)
    to_datetime = date_end - datetime.timedelta(
        days=1)
    interval1 = "5_MIN"  # ["DAY", "1_HR", "3_HR", "1_MIN", "5_MIN", "15_MIN", "60_MIN"]
    interval2 = "DAY"  # ["DAY", "1_HR", "3_HR", "1_MIN", "5_MIN", "15_MIN", "60_MIN"]
    indices = False
    df1 = pd.DataFrame(
        get_historical(instrument, from_datetime, to_datetime, interval1,
                       indices))

    df2 = pd.DataFrame(
        get_historical(instrument, from_datetime, to_datetime, interval2,
                       indices))

    q1 = """SELECT  volume FROM df1 order by volume DESC LIMIT 1 """
    q2 = """SELECT  open FROM df2 """
    q3 = """SELECT  close FROM df2 """

    s1 = ps.sqldf(q1, locals())
    s2 = ps.sqldf(q2, locals())
    s3 = ps.sqldf(q3, locals())

    vol = s1["volume"][0]
    op = s2["open"][0]
    clo = s3["close"][0]

    return vol, op, clo
