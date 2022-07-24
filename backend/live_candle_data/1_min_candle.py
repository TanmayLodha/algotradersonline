import datetime
import time
import threading
from alice_blue import *
import pandas as pd
from openpyxl import load_workbook, Workbook
from config1 import Credentials

# To make django-models work outside django
import sys
import os
import django

sys.path.append("/Users/nitishgupta/Desktop/algoTrade/backend")
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()
from optionchain.models import LTP

SCRIPT_LIST = [
    'ACC', 'ADANIENT', 'ADANIPORTS', 'AMBUJACEM', 'APOLLOHOSP', 'ASIANPAINT', 'AUBANK',
    'AUROPHARMA', 'AXISBANK', 'BAJAJ-AUTO', 'BAJFINANCE', 'BATAINDIA', 'BHARATFORG',
    'BHARTIARTL', 'BIOCON', 'BPCL', 'CHOLAFIN', 'CIPLA', 'COALINDIA', 'COFORGE', 'DABUR',
    'DIVISLAB', 'DLF', 'DRREDDY', 'EICHERMOT', 'GODREJCP', 'GODREJPROP', 'GRASIM', 'HAVELLS',
    'HCLTECH', 'HDFC', 'HDFCBANK', 'HDFCLIFE', 'HEROMOTOCO', 'HINDALCO', 'HINDPETRO',
    'HINDUNILVR', 'ICICIBANK', 'ICICIPRULI', 'IGL', 'INDIGO', 'INDUSINDBK', 'INFY',
    'IRCTC', 'ITC', 'JINDALSTEL', 'JSWSTEEL', 'JUBLFOOD', 'KOTAKBANK', 'LICHSGFIN',
    'LT', 'LTI', 'LUPIN', 'M&M', 'MARUTI', 'MINDTREE', 'MUTHOOTFIN', 'Nifty 50', 'Nifty Bank',
    'PEL', 'PIDILITIND', 'PVR', 'RELIANCE', 'SBICARD', 'SBILIFE', 'SBIN', 'SRF',
    'SRTRANSFIN', 'SUNPHARMA', 'TATACHEM', 'TATACONSUM', 'TATAMOTORS', 'TATAPOWER',
    'TATASTEEL', 'TCS', 'TECHM', 'TITAN', 'TVSMOTOR', 'UPL', 'VEDL', 'VOLTAS', 'WIPRO',
    'ZEEL']

socket_opened = False
df = pd.DataFrame()
df_final = pd.DataFrame()
ORB_timeFrame = 60  # in seconds
x = 1

# Creates a new file
filepath = f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}_1MIN.xlsx'
wb = Workbook()
wb.save(filepath)


def login():
    access_token = AliceBlue.login_and_get_access_token(
        username=Credentials.UserName.value,
        password=Credentials.PassWord.value,
        twoFA=Credentials.TwoFA.value,
        api_secret=Credentials.SecretKey.value,
        app_id=Credentials.AppId.value)
    alice = AliceBlue(username=Credentials.UserName.value,
                      password=Credentials.PassWord.value,
                      access_token=access_token)

    alice.start_websocket(subscribe_callback=event_handler_quote_update,
                          socket_open_callback=open_callback,
                          run_in_background=True)

    while not socket_opened:
        pass

    alice.subscribe([
        alice.get_instrument_by_symbol('NSE', i) for i in SCRIPT_LIST
    ], LiveFeedType.MARKET_DATA)

    alice.subscribe(
        alice.get_instrument_for_fno(symbol='BANKNIFTY', expiry_date=datetime.date(2022, 7, 28), is_fut=True,
                                     strike=None, is_CE=False), LiveFeedType.MARKET_DATA)
    alice.subscribe(
        alice.get_instrument_for_fno(symbol='NIFTY', expiry_date=datetime.date(2022, 7, 28), is_fut=True, strike=None,
                                     is_CE=False), LiveFeedType.MARKET_DATA)

    instruments_df = pd.DataFrame(alice.search_instruments('NFO', 'BANKNIFTY'))
    final_df = instruments_df.sort_values(by=['expiry'], ascending=True)

    q = LTP.objects.get(name='Nifty Bank')
    bn = q.ltp
    bn = int(bn / 100) * 100
    for i in range(bn - 1000, bn + 1000, 100):
        alice.subscribe(alice.get_instrument_for_fno(symbol='BANKNIFTY',
                                                     expiry_date=final_df.iloc[0]['expiry'],
                                                     is_fut=False,
                                                     strike=i,
                                                     is_CE=False),
                        LiveFeedType.MARKET_DATA)
        alice.subscribe(alice.get_instrument_for_fno(symbol='BANKNIFTY',
                                                     expiry_date=final_df.iloc[0]['expiry'],
                                                     is_fut=False,
                                                     strike=i,
                                                     is_CE=True),
                        LiveFeedType.MARKET_DATA)

    q = LTP.objects.get(name='Nifty 50')
    n = q.ltp
    n = int(n / 100) * 100
    for i in range(n - 1000, n + 1000, 100):
        alice.subscribe(alice.get_instrument_for_fno(symbol='NIFTY',
                                                     expiry_date=final_df.iloc[0]['expiry'],
                                                     is_fut=False,
                                                     strike=i,
                                                     is_CE=False),
                        LiveFeedType.MARKET_DATA)
        alice.subscribe(alice.get_instrument_for_fno(symbol='NIFTY',
                                                     expiry_date=final_df.iloc[0]['expiry'],
                                                     is_fut=False,
                                                     strike=i,
                                                     is_CE=True),
                        LiveFeedType.MARKET_DATA)


def event_handler_quote_update(message):
    ltp = message['ltp']
    instrument = message['instrument'].symbol

    # update live LTP to database
    # LTP.objects.create(name=instrument)
    # qt= LTP.objects.get(name=instrument)
    # qt.ltp = ltp
    # qt.save()

    p, created = LTP.objects.get_or_create(name=instrument)
    p.ltp = ltp
    p.save()

    global df
    if instrument in SCRIPT_LIST or instrument == "BANKNIFTY JUL FUT" or instrument == "NIFTY JUL FUT":
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        vol = message['volume']
        exchange = message['instrument'].exchange
        high = message['high']
        low = message['low']
        atp = message['atp']
        df_new = pd.DataFrame(
            {
                'symbol': instrument,
                'timestamp': timestamp,
                'vol': vol,
                'ltp': ltp,
                'high': high,
                'low': low,
                'exchange': exchange,
                'atp': atp
            },
            index=[0])
        df = pd.concat([df, df_new], ignore_index=True)


def open_callback():
    global socket_opened
    socket_opened = True
    print("Socket opened")


def create_ohlc():
    start = time.time()
    global df
    copydf = df.copy(deep=True).drop_duplicates()
    print(copydf)
    df = df.iloc[0:0]
    get_ohlc(copydf)
    interval = ORB_timeFrame - (time.time() - start)
    print(
        f"Next check will start after {interval} sec : {datetime.datetime.now()}"
    )

    threading.Timer(interval, create_ohlc).start()


def get_ohlc(dataframe):
    grouped = dataframe.groupby('symbol')

    global df_final
    global x
    book = load_workbook(
        f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}_1MIN.xlsx')
    writer = pd.ExcelWriter(
        f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}_1MIN.xlsx',
        engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    x = 1
    for name, group in grouped:
        group = group.sort_values('timestamp')
        timestamp = group['timestamp'].iloc[0]
        symbol = name
        volume = group['vol'].iloc[-1] - group['vol'].iloc[0]
        open = group['ltp'].iloc[0]
        close = group['ltp'].iloc[-1]
        high = group['ltp'].max()
        low = group['ltp'].min()
        exchange = group['exchange'].iloc[0]
        atp = group['atp'].iloc[-1]
        data = {
            'timestamp': timestamp,
            'symbol': symbol,
            'volume': volume,
            'open': open,
            'close': close,
            'high': high,
            'low': low,
            'exchange': exchange,
            'atp': atp
        }

        df_append = pd.DataFrame(data, index=[0])
        df_append.to_excel(writer, header=False, index=False, startrow=x, startcol=0)
        x += 1
    writer.save()
    book.close()


if __name__ == '__main__':
    while ((datetime.datetime.now().time() <= datetime.time(9, 14, 00))
           or (datetime.datetime.now().time() >= datetime.time(15, 30, 00))):
        pass

    login()
    main_interval = ORB_timeFrame - datetime.datetime.now().second

    print("start in ", main_interval)
    time.sleep(main_interval)
    create_ohlc()
