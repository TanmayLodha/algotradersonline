import datetime
import time
import threading
from alice_blue import *
import pandas as pd
from openpyxl import load_workbook, Workbook
from config import Credentials

SCRIPT_LIST = [
    'ACC', 'ADANIENT', 'ADANIPORTS', 'AMBUJACEM', 'APOLLOHOSP', 'ASIANPAINT', 'AUBANK',
    'AUROPHARMA', 'AXISBANK', 'BAJAJ-AUTO', 'BAJFINANCE', 'BATAINDIA', 'BHARATFORG',
    'BHARTIARTL', 'BIOCON', 'BPCL', 'CHOLAFIN', 'CIPLA', 'COALINDIA', 'COFORGE', 'DABUR',
    'DIVISLAB', 'DLF', 'DRREDDY', 'EICHERMOT', 'GODREJCP', 'GODREJPROP', 'GRASIM', 'HAVELLS',
    'HCLTECH', 'HDFC', 'HDFCBANK', 'HDFCLIFE', 'HEROMOTOCO', 'HINDALCO', 'HINDPETRO',
    'HINDUNILVR', 'ICICIBANK', 'ICICIPRULI', 'IGL', 'INDIGO', 'INDUSINDBK', 'INFY',
    'IRCTC', 'ITC', 'JINDALSTEL', 'JSWSTEEL', 'JUBLFOOD', 'KOTAKBANK', 'LICHSGFIN',
    'LT', 'LTI', 'LUPIN', 'M&M', 'MANAPPURAM', 'MARUTI', 'MINDTREE', 'MUTHOOTFIN',
    'PEL', 'PIDILITIND', 'PVR', 'RELIANCE', 'SBICARD', 'SBILIFE', 'SBIN', 'SRF',
    'SRTRANSFIN', 'SUNPHARMA', 'TATACHEM', 'TATACONSUM', 'TATAMOTORS', 'TATAPOWER',
    'TATASTEEL', 'TCS', 'TECHM', 'TITAN', 'TVSMOTOR', 'UPL', 'VEDL', 'VOLTAS', 'WIPRO',
    'ZEEL'
]

socket_opened = False
df = pd.DataFrame()
df_final = pd.DataFrame()
ORB_timeFrame = 300  # in seconds
x = 1

# Creates a new file
filepath = f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}.xlsx'
wb = Workbook()
wb.save(filepath)


def login():
    access_token = AliceBlue.login_and_get_access_token(
        username=Credentials.UserName.value,
        password=Credentials.PassWord.value,
        twoFA=Credentials.TwoFA.value,
        api_secret=Credentials.SecretKey.value,
        app_id=Credentials.AppId.value)
    alice = AliceBlue(username=Credentials.UserName.value,
                      password=Credentials.PassWord.value,
                      access_token=access_token)

    alice.start_websocket(subscribe_callback=event_handler_quote_update,
                          socket_open_callback=open_callback,
                          run_in_background=True)

    while not socket_opened:
        pass

    alice.subscribe([
        alice.get_instrument_by_symbol('NSE', i.upper()) for i in SCRIPT_LIST
    ], LiveFeedType.MARKET_DATA)


def event_handler_quote_update(message):
    # print(message)
    ltp = message['ltp']
    timestamp = pd.to_datetime(message['exchange_time_stamp'], unit='s')
    # timestamp = datetime.datetime.fromtimestamp(message['exchange_time_stamp'])
    vol = message['volume']
    instrument = message['instrument'].symbol
    exchange = message['instrument'].exchange
    high = message['high']
    low = message['low']
    global df
    # current_time = time.strftime("%Y-%m-%d %H:%M:%S", localtime())
    df_new = pd.DataFrame(
        {
            'symbol': instrument,
            'timestamp': timestamp,
            'vol': vol,
            'ltp': ltp,
            'high': high,
            'low': low,
            'exchange': exchange
        },
        index=[0])
    df = pd.concat([df, df_new], ignore_index=True)
    print(f"{instrument} : {ltp} : {timestamp} : {vol}")


def open_callback():
    global socket_opened
    socket_opened = True
    print("Socket opened")


def create_ohlc():
    start = time.time()
    global df
    copydf = df.copy(deep=True).drop_duplicates()
    df = pd.DataFrame()
    get_ohlc(copydf)
    interval = ORB_timeFrame - (time.time() - start)
    print(
        f"Next check will start after {interval} sec : {datetime.datetime.now()}"
    )

    threading.Timer(interval, create_ohlc).start()


def get_ohlc(dataframe):
    grouped = dataframe.groupby('symbol')
    global df_final
    global x
    book = load_workbook(
        f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}.xlsx')
    writer = pd.ExcelWriter(
        f'/Users/nitishgupta/Desktop/algoTrade/day_data/{datetime.datetime.now().strftime("%Y-%m-%d")}.xlsx',
        engine='openpyxl')
    writer.book = book

    for name, group in grouped:
        group = group.sort_values('timestamp')
        timestamp = group['timestamp'].iloc[0]
        symbol = name
        volume = group['vol'].iloc[-1] - group['vol'].iloc[0]
        open = group['ltp'].iloc[0]
        close = group['ltp'].iloc[-1]
        high = group['ltp'].max()
        low = group['ltp'].min()
        exchange = group['exchange'].iloc[0]
        data = {
            'timestamp': timestamp,
            'symbol': symbol,
            'volume': volume,
            'open': open,
            'close': close,
            'high': high,
            'low': low,
            'exchange': exchange
        }

        df_append = pd.DataFrame(data, index=[0])
        df_final = pd.concat([df_final, df_append], ignore_index=True)
        df_append.to_excel(writer, header=False, index=False, startrow=x)
        x += 1
    writer.save()
    book.close()


if __name__ == '__main__':
    while ((datetime.datetime.now().time() <= datetime.time(9, 14, 25))
           or (datetime.datetime.now().time() >= datetime.time(15, 30, 00))):
        pass

    login()
    # interval = ORB_timeFrame - datetime.datetime.now().second
    main_interval = (5 - datetime.datetime.now().minute % 5) * 60 - (
        datetime.datetime.now().second)
    print("start in ", main_interval)
    time.sleep(main_interval)
    create_ohlc()
